from openpyxl import Workbook
wb = Workbook()
mult_sheet = wb.active
mult_sheet.title = "Multiplication table"
n = int(input("Enter n for forming a N*N multplication table : "))
for i in range(2, n + 2):
    mult_sheet.cell(1, i).value = i - 1
    mult_sheet.cell(i, 1).value = i - 1
for i in range(2, n + 2):
    for j in range(2, n + 2):
        mult_sheet.cell(j, i).value = (i - 1) * (j - 1)
print("Multiplication table sucessfully created!")
wb.save("/Users/rishabh.kapoor/Desktop/Multiplication_Sheet.xlsx")